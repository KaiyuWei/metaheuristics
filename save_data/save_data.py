import pandas as pd
import os
from openpyxl import Workbook
from typing import List
from models.multiple_solution.evolutionary_based.tuner.Preset import Preset

def result_to_excel(df, file_name, s_name, i):
    exists = os.path.isfile(file_name)
    if not exists:
        wb = Workbook()
        wb.save(file_name)

    with pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df.to_excel(writer, sheet_name=s_name, startcol=(4 * i))

def presets_to_excel(lp: List[Preset], file_name, i):
    allPresets = pd.DataFrame()
    for pre in lp:
        df = pd.DataFrame(pre.parameters, index=[pre.name])
        allPresets = allPresets.append(df)
    
    exists = os.path.isfile(file_name)
    if not exists:
        wb = Workbook()
        wb.save(file_name)

    with pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        allPresets.to_excel(writer, sheet_name='Presets', startcol=(4 * i))

def save_probabilities(log_prob, file_name, i):
    """
    for saving probabilities after each update of presets probabilities
    """
    exists = os.path.isfile(file_name)
    if not exists:
        wb = Workbook()
        wb.save(file_name)
    
    with pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        log_prob.to_excel(writer, sheet_name=('probabilities exepriment {}'.format(str(i))))

